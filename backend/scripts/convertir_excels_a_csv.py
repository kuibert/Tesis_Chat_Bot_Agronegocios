"""
Script definitivo: Convierte archivos Excel del MAG a CSV limpios.
Usa openpyxl para leer correctamente celdas combinadas.
Procesa TODAS las hojas de calendario (las que tienen "Semana", "DDT", etc.).
"""

import openpyxl
import pandas as pd
from pathlib import Path
import re
import xlrd

# ---- CONFIGURACIÓN ----
SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR.parent.parent / "data" / "Calendarios de Fertilizacion-20260214T220614Z-1-001" / "Calendarios de Fertilizacion"
OUTPUT_DIR = SCRIPT_DIR.parent.parent / "data" / "csv"
HOJAS_IGNORADAS = ["main", "portada", "carátula", "índice"]
from excel_parser import (
    extraer_cabeceras_con_xlrd, 
    extraer_cabeceras_calendario, 
    hacer_cabeceras_unicas, 
    obtener_valor_celda_combinada,
    extraer_catalogo_fertilizantes,
    ADMIN_KEYWORDS
)

def es_hoja_ignorada(nombre_hoja):
    nombre_lower = nombre_hoja.lower()
    return any(p in nombre_lower for p in HOJAS_IGNORADAS)

def procesar_archivo(archivo_path):
    print(f"\n📄 Procesando: {archivo_path.name}")

    # Buscar el archivo .xls original para usar xlrd (más preciso que openpyxl para estos archivos)
    xls_original = archivo_path.with_suffix(".xls")
    usar_xlrd = xls_original.exists()
    if usar_xlrd:
        print(f"   ✅ Usando xlrd sobre archivo original: {xls_original.name}")
    else:
        print(f"   ⚠️  Archivo .xls original no encontrado, usando openpyxl como fallback.")

    wb_form = openpyxl.load_workbook(archivo_path, data_only=False)
    wb_val = openpyxl.load_workbook(archivo_path, data_only=True)

    for nombre_hoja in wb_form.sheetnames:
        if es_hoja_ignorada(nombre_hoja):
            continue
        ws_form = wb_form[nombre_hoja]
        ws_val = wb_val[nombre_hoja]

        # Hoja de catálogo de fertilizantes — NO es un calendario, se procesa aparte
        if nombre_hoja.lower().startswith("fertiliz") or "fertilizante" in nombre_hoja.lower():
            if not usar_xlrd:
                print(f"   ⚠️  Hoja '{nombre_hoja}': no hay .xls original disponible, se omite el catálogo (extraer_catalogo_fertilizantes requiere xlrd).")
                continue

            catalogo = extraer_catalogo_fertilizantes(xls_original, nombre_hoja)
            if not catalogo:
                continue

            df_catalogo = pd.DataFrame(catalogo)
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            nombre_csv = f"{archivo_path.stem}_Fertilizantes.csv"
            ruta_salida = OUTPUT_DIR / nombre_csv
            df_catalogo.to_csv(ruta_salida, index=False, encoding="utf-8")
            print(f"   ✅ {nombre_csv} ({len(df_catalogo)} fertilizantes)")
            continue

        # Intentar extracción de cabeceras con xlrd primero (más preciso)
        cabeceras = None
        fila_datos_0based = None
        if usar_xlrd:
            cabeceras, fila_datos_0based = extraer_cabeceras_con_xlrd(xls_original, nombre_hoja)

        # Si xlrd falló o no está disponible, usar openpyxl como fallback
        if cabeceras is None:
            cabeceras_openpyxl, fila_datos_1based = extraer_cabeceras_calendario(ws_form, ws_val, wb_form, wb_val)
            if cabeceras_openpyxl is None:
                print(f"   ⚠️  Hoja '{nombre_hoja}' no tiene cabecera de calendario, se omite.")
                continue
            cabeceras = cabeceras_openpyxl
            fila_datos_0based = fila_datos_1based - 1  # convertir a 0-based para skiprows

        cabeceras = hacer_cabeceras_unicas(cabeceras)

        # Leer datos con pandas a partir de la fila correcta (skiprows es 0-based)
        df = pd.read_excel(archivo_path, sheet_name=nombre_hoja, header=None, skiprows=fila_datos_0based)
        df.columns = cabeceras[:len(df.columns)]  # Ajustar por si acaso

        # Marcar columnas administrativas
        nuevas_cols = []
        for col in df.columns:
            col_lower = str(col).lower()
            if any(kw in col_lower for kw in ADMIN_KEYWORDS):
                nuevas_cols.append(f"Admin_{col}")
            else:
                nuevas_cols.append(col)
        df.columns = nuevas_cols

        # Encontrar la columna de semana
        col_semana = [col for col in df.columns if 'semana' in str(col).lower()]
        if col_semana:
            col_name = col_semana[0]
            # Si es un DataFrame por duplicados, obtener la primera columna Series
            series_semana = df[col_name]
            if isinstance(series_semana, pd.DataFrame):
                series_semana = series_semana.iloc[:, 0]
                
            df[col_name] = pd.to_numeric(series_semana, errors='coerce')
            df = df.dropna(subset=[col_name])
            
            # Volver a filtrar de manera segura
            series_filtrado = df[col_name]
            if isinstance(series_filtrado, pd.DataFrame):
                series_filtrado = series_filtrado.iloc[:, 0]
            df = df[series_filtrado > 0]
            
            # Castear a entero de manera segura
            series_final = df[col_name]
            if isinstance(series_final, pd.DataFrame):
                series_final = series_final.iloc[:, 0]
            df[col_name] = series_final.astype(int)

        if df.empty:
            print(f"   ⚠️  Hoja '{nombre_hoja}' sin datos válidos, se omite.")
            continue

        # Guardar CSV
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        nombre_csv = f"{archivo_path.stem}_{nombre_hoja.replace(' ', '_')}.csv"
        ruta_salida = OUTPUT_DIR / nombre_csv
        df.to_csv(ruta_salida, index=False, encoding='utf-8')
        print(f"   ✅ {nombre_csv} ({len(df)} filas)")

    # Cerrar workbooks y eliminar el archivo temporal xlsx si vino de un .xls
    wb_form.close()
    wb_val.close()
    original_xls = archivo_path.with_suffix(".xls")
    if original_xls.exists():
        try:
            archivo_path.unlink()
        except Exception:
            pass

def main():
    print("🚀 Iniciando proceso de conversión...")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Buscar los archivos .xlsx listos para procesar
    archivos = list(INPUT_DIR.glob("*.xlsx"))
    if not archivos:
        print(f"❌ No se encontraron archivos Excel en '{INPUT_DIR}'")
        return

    for archivo in archivos:
        procesar_archivo(archivo)

    print(f"\n✨ Conversión completada. Los CSV están en '{OUTPUT_DIR}'")

if __name__ == "__main__":
    main()
